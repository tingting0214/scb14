# -*- coding: utf-8 -*-
# @Time ： 2020/7/15 14:01
# @Auth ： Carrie
# @File ：lessons7.py
# @QQ ：654291275
import requests
import openpyxl  #导入库
#读取测试用例函数
def read_data(filename,sheetname):
    wb=openpyxl.load_workbook(filename)  #加载工作簿---填入文档名字
    sheet=wb[sheetname]                   #获取表单
    max_row=sheet.max_row     #获取最大行数
    case_list=[]   #创建空列表，存放测试用例
    for i in range(2,max_row+1):
       dict1=dict(
       case_id=sheet.cell(row=i,column=1).value,   #通过表单获取行号，列号，通过.value来提取数据
       url=sheet.cell(row=i,column=5).value,      #获取url
       data=sheet.cell(row=i,column=6).value,      #获取data
       expect=sheet.cell(row=i,column=7).value    #获取expect
       )
       case_list.append(dict1)    #每循环一次，就把读取到的字典数据存放到这个list里
    return case_list   #返回测试用例列表
#执行接口函数
def api_fun(url,data):
    heaers_rec = {"X-Lemonban-Media-Type": "lemonban.v2","Content-Type": "application/json" }
    res=requests.post(url=url,json=data,headers=heaers_rec)
    respon=res.json()
    return respon
#写入结果
def write_result(filename,sheetname,row,column,final_result):
  wb=openpyxl.load_workbook(filename)
  sheet=wb[sheetname]
  sheet.cell(row=row,column=column).value=final_result   #写人结果
  wb.save(filename)     #保存，要先把表格关闭，才能保存


def execult(filename,sheetname):
    cares = read_data(filename,sheetname)  #调用读取测试用例，获取所有测试用例数据保存到变量
    for case in cares:   #遍历里面的每一个字典
        case_id= case.get("case_id")  #case["case_id}  字典型，所有写key值，取value值
        url=case.get("url")
        data=eval(case.get("data"))      # eval（）--运行被字符串包裹的表达式
        expect=eval(case.get("expect"))       #获取预期结果
        expect_msg=expect.get("msg")     #获取预期结果中的msg
        real_resule=api_fun(url=url,data=data)   #调用发送接口请求函数，返回结果用变量接收
        real_msg=real_resule.get("msg")   #获取实际结果中的msg
        print("预期结果中的msg：{}".format(expect_msg))
        print("实际结果中的msg：{}".format(real_msg))
        if real_msg==expect_msg:
            print("这{}条用例执行通过！".format(case_id))
            final_re="Passed"
        else:
            print("这{}条用例执行不通过！".format(case_id))
            final_re="Failed"
        write_result(filename,sheetname,case_id+1,8,final_re)
        print("*"*25)
execult("test_case_api.xlsx","login")


